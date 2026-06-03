from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from typing import Optional
from datetime import datetime, date
import io
from app.database import get_db
from app.dependencies import require_admin
from app import models

router = APIRouter(prefix="/api/reports", tags=["reports"])

def get_visits_query(db, user_id, checkpoint_id, date_from, date_to):
    q = db.query(models.VisitRecord).options(
        joinedload(models.VisitRecord.user),
        joinedload(models.VisitRecord.checkpoint)
    )
    if user_id:
        q = q.filter(models.VisitRecord.user_id == user_id)
    if checkpoint_id:
        q = q.filter(models.VisitRecord.checkpoint_id == checkpoint_id)
    if date_from:
        q = q.filter(models.VisitRecord.visited_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.filter(models.VisitRecord.visited_at <= datetime.combine(date_to, datetime.max.time()))
    return q.order_by(models.VisitRecord.visited_at.desc())

@router.get("/excel")
def export_excel(
    user_id: Optional[int] = None,
    checkpoint_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _=Depends(require_admin)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    records = get_visits_query(db, user_id, checkpoint_id, date_from, date_to).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش بازدیدها"
    ws.sheet_view.rightToLeft = True

    headers = ["ردیف", "کارشناس", "نقطه بازرسی", "محل", "تاریخ و ساعت", "وضعیت", "توضیحات", "دستگاه"]
    header_fill = PatternFill("solid", fgColor="1E40AF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_map = {"ok": "سالم", "issue": "مشکل دار", "critical": "بحرانی"}
    for i, rec in enumerate(records, 1):
        ws.append([
            i,
            rec.user.full_name if rec.user else "",
            rec.checkpoint.name if rec.checkpoint else "",
            rec.checkpoint.location if rec.checkpoint else "",
            rec.visited_at.strftime("%Y-%m-%d %H:%M"),
            status_map.get(rec.status.value, rec.status.value),
            rec.notes or "",
            rec.device_id or "",
        ])
        if rec.status.value == "critical":
            for col in range(1, 9):
                ws.cell(row=i+1, column=col).fill = PatternFill("solid", fgColor="FEE2E2")
        elif rec.status.value == "issue":
            for col in range(1, 9):
                ws.cell(row=i+1, column=col).fill = PatternFill("solid", fgColor="FEF9C3")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=visits_report.xlsx"}
    )

@router.get("/pdf")
def export_pdf(
    user_id: Optional[int] = None,
    checkpoint_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _=Depends(require_admin)
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    records = get_visits_query(db, user_id, checkpoint_id, date_from, date_to).all()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm)

    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph("NFC Attendance Report", styles["Title"]))
    elements.append(Spacer(1, 0.5*cm))

    data = [["#", "Expert", "Checkpoint", "Location", "Date/Time", "Status", "Notes"]]
    status_map = {"ok": "OK", "issue": "Issue", "critical": "Critical"}
    for i, rec in enumerate(records, 1):
        data.append([
            str(i),
            rec.user.full_name if rec.user else "",
            rec.checkpoint.name if rec.checkpoint else "",
            rec.checkpoint.location if rec.checkpoint else "",
            rec.visited_at.strftime("%Y-%m-%d %H:%M"),
            status_map.get(rec.status.value, rec.status.value),
            (rec.notes or "")[:40],
        ])

    col_widths = [1*cm, 4*cm, 4*cm, 4*cm, 4*cm, 2.5*cm, 6*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("PADDING", (0,0), (-1,-1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=visits_report.pdf"})

@router.get("/summary")
def summary_stats(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _=Depends(require_admin)
):
    q = db.query(models.VisitRecord)
    if date_from:
        q = q.filter(models.VisitRecord.visited_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.filter(models.VisitRecord.visited_at <= datetime.combine(date_to, datetime.max.time()))

    by_user = db.query(
        models.User.full_name,
        func.count(models.VisitRecord.id).label("count")
    ).join(models.VisitRecord, models.User.id == models.VisitRecord.user_id).group_by(models.User.full_name).all()

    by_checkpoint = db.query(
        models.Checkpoint.name,
        func.count(models.VisitRecord.id).label("count")
    ).join(models.VisitRecord, models.Checkpoint.id == models.VisitRecord.checkpoint_id).group_by(models.Checkpoint.name).all()

    by_status = db.query(
        models.VisitRecord.status,
        func.count(models.VisitRecord.id).label("count")
    ).group_by(models.VisitRecord.status).all()

    return {
        "by_user": [{"name": r[0], "count": r[1]} for r in by_user],
        "by_checkpoint": [{"name": r[0], "count": r[1]} for r in by_checkpoint],
        "by_status": [{"status": r[0].value, "count": r[1]} for r in by_status],
    }
